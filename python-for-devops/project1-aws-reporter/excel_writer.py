import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def write_to_excel(filename, sheets):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

    for sheet_name, (headers, rows) in sheets.items():
        ws = wb.create_sheet(title=sheet_name)

        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_num, row_data in enumerate(rows, start=2):
            for col_num, value in enumerate(row_data, start=1):
                ws.cell(row=row_num, column=col_num, value=value)

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    wb.save(filename)
